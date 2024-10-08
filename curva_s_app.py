import pandas as pd
import matplotlib.pyplot as plt
import streamlit as st
import io
from openpyxl import Workbook
from openpyxl.utils.dataframe import dataframe_to_rows
from openpyxl.chart import LineChart, Reference
from fpdf import FPDF
import tempfile
import os

# Função para limpar a abreviação dos dias da semana
def clean_weekday_abbreviation(date_str):
    return date_str.split(' ', 1)[1] if isinstance(date_str, str) else date_str

# Função para ler o arquivo Excel e tratar as colunas de data
def read_excel(file):
    df = pd.read_excel(file)
    
    # Limpar as colunas de datas
    df['Início'] = df['Início'].apply(lambda x: clean_weekday_abbreviation(x) if isinstance(x, str) else x)
    df['Término'] = df['Término'].apply(lambda x: clean_weekday_abbreviation(x) if isinstance(x, str) else x)
    
    # Converter para datetime
    df['Início'] = pd.to_datetime(df['Início'], format='%d/%m/%y', errors='coerce')
    df['Término'] = pd.to_datetime(df['Término'], format='%d/%m/%y', errors='coerce')
    
    # Tratar a duração (remover "dias" e converter para float)
    if 'Duração' in df.columns:
        df['Duracao'] = df['Duração'].str.extract('(\d+)').astype(float)
    
    return df

# Função para gerar a Curva S
def gerar_curva_s(df_raw, start_date_str='16/09/2024'):
    df_raw['Início'] = df_raw['Início'].apply(lambda x: clean_weekday_abbreviation(x) if isinstance(x, str) else x)
    df_raw['Término'] = df_raw['Término'].apply(lambda x: clean_weekday_abbreviation(x) if isinstance(x, str) else x)
    
    start_date = pd.to_datetime(start_date_str)
    end_date = df_raw['Término'].max()
    weeks = pd.date_range(start=start_date, end=end_date, freq='W-MON')

    progress_by_week = pd.DataFrame(weeks, columns=['Data'])
    progress_by_week['% Executado'] = 0.0

    for i, row in df_raw.iterrows():
        if pd.notna(row['Início']) and pd.notna(row['Término']):
            task_weeks = pd.date_range(start=row['Início'], end=row['Término'], freq='W-MON')
            if len(task_weeks) == 0:
                weekly_progress = 1  # Se a tarefa durar menos de uma semana
                week = row['Início']
                progress_by_week.loc[progress_by_week['Data'] == week, '% Executado'] += weekly_progress
            else:
                weekly_progress = 1 / len(task_weeks)
                for week in task_weeks:
                    progress_by_week.loc[progress_by_week['Data'] == week, '% Executado'] += weekly_progress

    progress_by_week['% Executado Acumulado'] = progress_by_week['% Executado'].cumsum() * 100

    max_progress = progress_by_week['% Executado Acumulado'].max()
    if max_progress > 0:
        progress_by_week['% Executado Acumulado'] = (progress_by_week['% Executado Acumulado'] / max_progress) * 100

    # Adicionar Delta entre as semanas (diferença entre a semana atual e a anterior)
    progress_by_week['Delta'] = progress_by_week['% Executado Acumulado'].diff().fillna(0)

    # Plotar a Curva S e retornar o gráfico como imagem
    fig, ax = plt.subplots(figsize=(10, 6))
    ax.plot(progress_by_week['Data'], progress_by_week['% Executado Acumulado'], marker='o', linestyle='-', color='b')
    ax.set_title('Curva S - % Executado por Semana')
    ax.set_xlabel('Data')
    ax.set_ylabel('% Executado Acumulado')
    ax.grid(True)
    plt.xticks(rotation=45)
    plt.tight_layout()

    # Salvar a imagem como um arquivo temporário
    with tempfile.NamedTemporaryFile(delete=False, suffix='.png') as temp_img_file:
        curva_s_img_path = temp_img_file.name
        fig.savefig(curva_s_img_path, format='png')

    plt.close(fig)

    st.pyplot(fig)

    return progress_by_week, curva_s_img_path

# Função para exportar o relatório para Excel com 3 abas
def export_to_excel(df, curva_s_df, atividades_proxima_semana, atividades_proximos_15_dias):
    output = io.BytesIO()
    
    wb = Workbook()

    # Aba Curva S
    ws1 = wb.active
    ws1.title = 'Curva S'
    
    for r in dataframe_to_rows(curva_s_df, index=False, header=True):
        ws1.append(r)
    
    chart = LineChart()
    chart.title = "Curva S - Progresso Acumulado"
    chart.y_axis.title = 'Progresso Acumulado (%)'
    chart.x_axis.title = 'Data'
    
    data = Reference(ws1, min_col=2, min_row=2, max_row=len(curva_s_df) + 1, max_col=2)
    chart.add_data(data, titles_from_data=True)
    
    ws1.add_chart(chart, "E5")

    # Aba Atividades para Próxima Semana
    ws2 = wb.create_sheet(title="Atividades Próxima Semana")
    for r in dataframe_to_rows(atividades_proxima_semana, index=False, header=True):
        ws2.append(r)

    # Aba Atividades para Próximos 15 Dias
    ws3 = wb.create_sheet(title="Atividades Próximos 15 Dias")
    for r in dataframe_to_rows(atividades_proximos_15_dias, index=False, header=True):
        ws3.append(r)
    
    wb.save(output)
    output.seek(0)
    
    return output

# Função para gerar o relatório em PDF
def gerar_relatorio_pdf(df, caminho_critico, atividades_sem_predecessora, atividades_atrasadas, curva_s_img_path):
    pdf = FPDF()

    pdf.add_page()
    pdf.set_font("Arial", size=12)

    # Título do relatório
    pdf.cell(200, 10, txt="Relatório Detalhado do Projeto", ln=True, align="C")
    pdf.ln(10)  # Espaço entre título e próximo conteúdo

    # Adicionar Curva S
    pdf.cell(200, 10, txt="Curva S", ln=True)
    pdf.image(curva_s_img_path, x=10, y=40, w=190)  # Ajustar a posição e tamanho da imagem
    pdf.ln(175)  # Adicionar espaçamento abaixo da imagem

    # Adicionar caminho crítico
    pdf.cell(200, 10, txt="Caminho Crítico", ln=True)
    pdf.ln(5)  # Espaçamento
    for atividade in caminho_critico:
        pdf.cell(200, 10, txt=atividade, ln=True)
    pdf.ln(10)  # Espaçamento entre seções

    # Adicionar atividades sem predecessoras
    pdf.cell(200, 10, txt="Atividades Sem Predecessoras", ln=True)
    atividades_sem_predecessora_df = pd.DataFrame(atividades_sem_predecessora)
    pdf.ln(5)  # Espaçamento
    for _, row in atividades_sem_predecessora_df.iterrows():
        pdf.cell(200, 10, txt=row['Nome da tarefa'], ln=True)
    pdf.ln(10)  # Espaçamento entre seções

    # Adicionar atividades atrasadas
    if not atividades_atrasadas.empty:
        pdf.cell(200, 10, txt="Atividades Atrasadas", ln=True)
        pdf.ln(5)  # Espaçamento
        for _, row in atividades_atrasadas.iterrows():
            pdf.cell(200, 10, txt=row['Nome da tarefa'], ln=True)
    pdf.ln(10)  # Espaçamento final

    # Salvar o relatório em PDF no objeto BytesIO
    pdf_output = io.BytesIO()
    pdf_output.write(pdf.output(dest='S').encode('latin1'))  # Salva diretamente no fluxo de bytes
    pdf_output.seek(0)

    # Remover o arquivo temporário de gráfico
    if os.path.exists(curva_s_img_path):
        os.remove(curva_s_img_path)

    return pdf_output

# Interface Streamlit
st.title('AWPlan - A ferramenta de Gestão de Cronograma')

uploaded_file = st.file_uploader("Escolha o arquivo Excel do cronograma", type="xlsx")

start_date = st.text_input("Selecione a data de início do projeto (DD/MM/AAAA)", placeholder="DD/MM/AAAA")

if st.button("Gerar Relatório"):
    if uploaded_file is not None and start_date:
        try:
            # Carregar o Excel
            df_raw = read_excel(uploaded_file)

            # Gerar Curva S e obter o gráfico como imagem
            progress_by_week, curva_s_img_path = gerar_curva_s(df_raw, start_date_str=start_date)

            # Abas para visualização com botões expansíveis
            with st.expander(" ▶️ Dados do Cronograma"):
                st.dataframe(df_raw)

            atividades_sem_predecessora = df_raw[df_raw['Predecessoras'].isna()]

            with st.expander(" ▶️ Atividades sem Predecessoras"):
                st.dataframe(atividades_sem_predecessora)

            caminho_critico = df_raw[df_raw['Duracao'] > 15]  # Exemplo de caminho crítico simplificado
            with st.expander(" ▶️ Caminho Crítico"):
                st.dataframe(caminho_critico)

            atividades_atrasadas = df_raw[df_raw['Término'] < pd.Timestamp.today()]
            with st.expander(" ▶️ Atividades Atrasadas"):
                st.dataframe(atividades_atrasadas)

            proximos_7_dias = pd.Timestamp.today() + pd.Timedelta(days=7)
            atividades_proxima_semana = df_raw[(df_raw['Início'] <= proximos_7_dias) & (df_raw['Término'] >= pd.Timestamp.today())]

            with st.expander(" ▶️ Atividades para Próxima Semana"):
                st.dataframe(atividades_proxima_semana)

            proximos_15_dias = pd.Timestamp.today() + pd.Timedelta(days=15)
            atividades_proximos_15_dias = df_raw[(df_raw['Início'] <= proximos_15_dias) & (df_raw['Término'] >= pd.Timestamp.today())]
            
            with st.expander("  ▶️ Atividades para os Próximos 15 Dias"):
                st.dataframe(atividades_proximos_15_dias)

            # Gerar Relatório em PDF com a imagem da Curva S
            pdf_data = gerar_relatorio_pdf(df_raw, caminho_critico, atividades_sem_predecessora, atividades_atrasadas, curva_s_img_path)

            # Botão para baixar o relatório em PDF
            st.download_button(
                label="Baixar Relatório Gerencial em PDF",
                data=pdf_data.getvalue(),
                file_name="relatorio_projeto.pdf",
                mime="application/pdf"
            )

            # Exportar o Excel com as 3 abas
            excel_data = export_to_excel(df_raw, progress_by_week, atividades_proxima_semana, atividades_proximos_15_dias)

            # Botão para baixar o arquivo Excel
            st.download_button(
                label="Baixar Arquivo Excel com Curva S, Atividades da Semana e Atividades para os Próximos 15 dias",
                data=excel_data.getvalue(),
                file_name="cronograma_com_curva_s.xlsx",
                mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
            )

        except ValueError as e:
            st.error(f"Erro ao processar os dados: {e}")
